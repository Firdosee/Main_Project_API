import string
import random
import openpyxl

from Logger.Base_Logger import abc_test_Base

log = abc_test_Base.getLogger()


class Excel_Data:

    def getAPIData(self, data_section, testcase_name):
        try:
            data_dict = {}
            # log.info("a")
            book = openpyxl.load_workbook("../Data_Config/API_Data.xlsx")
            # log.info("b")
            # sheet = book.active
            sheet = book.get_sheet_by_name(data_section)
            # log.info("c")
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == testcase_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return data_dict
        except FileNotFoundError as e:
            print("Exception occurred while performing file operation", e)

    def gettaskData(self, test_name):
        try:
            data_dict = {}
            book = openpyxl.load_workbook("API_Data.xlsx")
            sheet = book.active
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == test_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        data_dict[sheet.cell(row=11, column=j).value] = sheet.cell(row=i, column=j).value
            return data_dict
        except FileNotFoundError as e:
            print("Exception occurred while performing file operation", e)

    def random_text_generator(self):
        try:
            letters = string.ascii_letters
            text = (''.join(random.choice(letters) for i in range(5)))
            print("Random generated text is ", text)
            return text
        except Exception as e:
            print("Exception occurred while creating generation of random string", e)




